"""xlsx <-> JSON read/write for the tk-vn product sheet.

Two subcommands:
  dump   <xlsx> <out.json>            read sheet, emit per-row JSON
  apply  <xlsx> <updates.json> <out>  write updates back into a copy of the xlsx

The sheet is treated as a single tab named `tiktok_chanpin_`. All columns are
preserved; only cells named in the updates file are overwritten.
"""
from __future__ import annotations

import json
import re
import shutil
import sys
from html import unescape
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

SHEET_NAME = "tiktok_chanpin_"

# Default column letters (used when header-based detection fails).
# Real TikTok exports may have different column orders, so we also support
# dynamic header-based lookup via `resolve_columns()`.
COL = {
    "cat_id": "A", "title": "B", "desc": "C", "brand": "D", "attrs": "E",
    "sku": "F", "vname1": "G", "vval1": "H", "vname2": "I", "vval2": "J",
    "vname3": "K", "vval3": "L", "stock": "Q",
    "main_img": "R",
    "sub_imgs": ["S", "T", "U", "V", "W", "X", "Y", "Z"],  # 附图一~八
    "video": "AA", "size_img": "AB", "variant_img": "AC",
    "weight": "AD", "length": "AE", "width": "AF", "height": "AG",
}

# Header name patterns for dynamic column detection.
# Keys match COL keys; values are substrings to match against row-1 headers.
HEADER_PATTERNS = {
    "title": ["产品标题", "标题"],
    "desc": ["Tiktok产品描述", "产品描述", "描述"],
    "brand": ["品牌"],
    "sku": ["sku", "SKU"],
    "vname1": ["变种属性名称一", "变种属性名称1"],
    "vval1": ["变种属性值一", "变种属性值1"],
    "vname2": ["变种属性名称二", "变种属性名称2"],
    "vval2": ["变种属性值二", "变种属性值2"],
    "vname3": ["变种属性名称三", "变种属性名称3"],
    "vval3": ["变种属性值三", "变种属性值3"],
    "stock": ["库存"],
    "main_img": ["主图", "主图(url)地址"],
    "video": ["视频连接", "视频链接", "视频"],
    "variant_img": ["变种主题1图片", "变种主题一图片", "变种图片"],
    "weight": ["重量(kg)", "重量"],
    "length": ["长", "长(cm)"],
    "width": ["宽", "宽(cm)"],
    "height": ["高", "高(cm)"],
    "price": ["价格(站点币种)", "价格（站点币种）", "价格"],
    "local_price": ["本地展示价"],
}

IMG_SRC_RE = re.compile(r'<img[^>]+src="([^"]+)"', re.IGNORECASE)


def col_idx(letter: str) -> int:
    s = 0
    for ch in letter:
        s = s * 26 + (ord(ch.upper()) - 64)
    return s


def resolve_columns(ws) -> dict:
    """Build a COL dict by matching HEADER_PATTERNS against ws row-1 headers.
    Falls back to the static COL for keys not found. Also detects 附图 columns
    by matching '附图' prefix. Returns dict with same keys as COL but letters
    adjusted to the actual sheet. Keys not found get empty string "".
    """
    from openpyxl.utils import get_column_letter
    headers = {}
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(row=1, column=c).value or "").strip()
        if h:
            headers[h] = get_column_letter(c)

    # Start with empty defaults — only fill keys we actually find
    resolved = {k: "" for k in COL}
    # Also include price/local_price keys from HEADER_PATTERNS
    for k in HEADER_PATTERNS:
        if k not in resolved:
            resolved[k] = ""

    # Map by header patterns: exact match first, then substring match
    for key, patterns in HEADER_PATTERNS.items():
        # Pass 1: exact match (h == p)
        for h, letter in headers.items():
            for p in patterns:
                if h == p:
                    resolved[key] = letter
                    break
            if resolved[key]:
                break
        if resolved[key]:
            continue
        # Pass 2: substring match (p in h) — lower priority
        for h, letter in headers.items():
            for p in patterns:
                if p in h:
                    resolved[key] = letter
                    break
            if resolved[key]:
                break

    # Detect 附图一~八 columns dynamically
    sub_imgs = []
    for h, letter in headers.items():
        if h.startswith("附图"):
            sub_imgs.append(letter)
    if sub_imgs:
        sub_imgs.sort(key=col_idx)
        resolved["sub_imgs"] = sub_imgs

    return resolved


def normalize_url(url: str | None) -> str:
    """Fix protocol-relative URLs (//img...) -> https://img...
    Returns empty string for non-URL values (numeric 0.3, 'nan', etc.)."""
    if not url:
        return ""
    u = str(url).strip()
    # Filter out non-URL values (numeric weights, 'nan', 'None', '0')
    if u in {"0", "0.0", "nan", "None", "0.3"}:
        return ""
    try:
        float(u)  # catches "0.3", "0.0", numeric strings
        return ""
    except (ValueError, TypeError):
        pass
    if u.startswith("//"):
        return "https:" + u
    if u.startswith("http://"):
        return "https://" + u[len("http://"):]
    if not u.startswith("https://"):
        return ""  # not a valid URL
    return u


def extract_img_urls(html: str | None) -> list[str]:
    if not html:
        return []
    seen: set[str] = set()
    urls: list[str] = []
    for m in IMG_SRC_RE.findall(html):
        u = normalize_url(unescape(m))
        if u and u not in seen:
            seen.add(u)
            urls.append(u)
    return urls


def build_description_html(urls: list[str]) -> str:
    return "".join(f'<img src="{u}">' for u in urls if u)


def extract_text_content(html: str | None) -> str:
    """Strip all <img ...> tags from HTML, return the remaining text HTML."""
    if not html:
        return ""
    # 删完整<img ...>标签(不只是src部分),避免残留 > 符号
    text = re.sub(r'<img[^>]*/?>', '', str(html), flags=re.IGNORECASE)
    return text.strip()


def build_description_all(text_html: str, img_urls: list[str]) -> str:
    """Combine translated text HTML + cleaned img URLs back into a description HTML.
    Text goes before images (matches actual data patterns)."""
    cleaned_imgs = "".join(f'<img src="{u}">' for u in img_urls if u)
    if text_html and cleaned_imgs:
        return text_html + cleaned_imgs
    return text_html or cleaned_imgs


def dump(xlsx_path: str, out_json: str) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    max_col = ws.max_column
    headers = {
        get_column_letter(c): ws.cell(row=1, column=c).value
        for c in range(1, max_col + 1)
    }
    rows = []
    for r in range(2, ws.max_row + 1):
        # skip fully empty rows
        if all((ws.cell(row=r, column=c).value in (None, "")) for c in range(1, max_col + 1)):
            continue
        fields = {
            get_column_letter(c): ws.cell(row=r, column=c).value
            for c in range(1, max_col + 1)
        }
        rows.append({"row_index": r, "fields": fields})
    out = {
        "sheet": ws.title,
        "max_col": max_col,
        "headers": headers,
        "rows": rows,
    }
    Path(out_json).write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"dumped {len(rows)} rows -> {out_json}")


def apply(xlsx_path: str, updates_json: str, out_xlsx: str) -> None:
    updates = json.loads(Path(updates_json).read_text(encoding="utf-8"))
    # updates: {"<row_index>": {"B": "...", ...}, ...}  (row_index as str key)
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    written = 0
    for row_key, cell_updates in updates.items():
        r = int(row_key)
        for letter, value in cell_updates.items():
            ws.cell(row=r, column=col_idx(letter)).value = value
            written += 1
    # If overwriting the source, back it up first.
    if Path(out_xlsx).resolve() == Path(xlsx_path).resolve():
        bak = Path(xlsx_path).with_suffix(xlsx_path.suffix + ".bak")
        shutil.copy2(xlsx_path, bak)
        print(f"backed up original -> {bak}")
    wb.save(out_xlsx)
    print(f"wrote {written} cells -> {out_xlsx}")


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print(__doc__)
        return 2
    cmd = argv[1]
    if cmd == "dump":
        dump(argv[2], argv[3])
    elif cmd == "apply":
        apply(argv[2], argv[3], argv[4])
    else:
        print(f"unknown command: {cmd}")
        return 2
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
