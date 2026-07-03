"""Sheet-level checks used by the eval runner.

Usage: python check_sheet.py <xlsx> <check_name>

check_name ∈ {brand_set, stock_set, video_cleared, sku_format,
image_urls_https, duplicate_images, required_fields, final_integrity}
Exit 0 if the check passes, 1 otherwise. Prints a short reason on failure.
"""
from __future__ import annotations

import re
import sys

import openpyxl

SHEET = "tiktok_chanpin_"


def _norm_header(value) -> str:
    text = re.sub(r"[\s*（）()]", "", str(value or "")).lower()
    return text.replace("必填", "")


def header_map(ws) -> dict[str, int]:
    return {_norm_header(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}


def find_col(ws, *needles: str) -> int | None:
    headers = header_map(ws)
    for needle in needles:
        n = _norm_header(needle)
        for header, col in headers.items():
            if n == header or n in header:
                return col
    return None


def data_rows(ws) -> list[int]:
    return [
        r for r in range(2, ws.max_row + 1)
        if any(ws.cell(r, c).value not in (None, "") for c in range(1, ws.max_column + 1))
    ]


def check_brand_set(ws) -> bool:
    col = find_col(ws, "品牌")
    return bool(col) and all(not ws.cell(r, col).value for r in data_rows(ws))


def check_stock_set(ws) -> bool:
    col = find_col(ws, "库存")
    return bool(col) and all(ws.cell(r, col).value == 30 for r in data_rows(ws))


def check_video_cleared(ws) -> bool:
    col = find_col(ws, "视频链接", "视频连接")
    return bool(col) and all(not ws.cell(r, col).value for r in data_rows(ws))


def check_sku_format(ws) -> bool:
    vals = [ws.cell(r, 6).value for r in data_rows(ws)]
    if not all(isinstance(v, str) and re.fullmatch(r"\d{13}", v) for v in vals):
        return False
    return vals == sorted(vals)


def check_image_urls_https(ws) -> bool:
    img_cols = [
        find_col(ws, "产品主图", "主图url地址"),
        *[find_col(ws, f"附图{x}") for x in "一二三四五六七八"],
        find_col(ws, "变种主题1图片"),
    ]
    img_cols = [c for c in img_cols if c]
    for r in data_rows(ws):
        for c in img_cols:
            v = ws.cell(r, c).value
            if v and not str(v).startswith("https://"):
                return False
        desc = ws.cell(r, 3).value or ""
        for u in re.findall(r'<img[^>]+src="([^"]+)"', desc):
            if not u.startswith("https://"):
                return False
    return True


def check_duplicate_images(ws) -> bool:
    cols = [
        find_col(ws, "产品主图", "主图url地址"),
        *[find_col(ws, f"附图{x}") for x in "一二三四五六七八"],
    ]
    cols = [c for c in cols if c]
    for r in data_rows(ws):
        values = [str(ws.cell(r, c).value).strip().lower() for c in cols if ws.cell(r, c).value]
        if len(values) != len(set(values)):
            return False
    return True


def check_required_fields(ws) -> bool:
    required = [
        ("分类id",), ("产品标题",), ("产品描述",), ("本地展示价",), ("库存",),
        ("产品主图",), ("重量kg",), ("长cm",), ("宽cm",), ("高cm",), ("仓库名称",),
    ]
    cols = [find_col(ws, *names) for names in required]
    if any(c is None for c in cols):
        return False
    title_col = find_col(ws, "产品标题")
    numeric_cols = {find_col(ws, "重量kg"), find_col(ws, "长cm"), find_col(ws, "宽cm"), find_col(ws, "高cm")}
    for r in data_rows(ws):
        for c in cols:
            value = ws.cell(r, c).value
            if value is None or str(value).strip() == "":
                return False
            if c in numeric_cols:
                try:
                    if float(value) <= 0:
                        return False
                except (TypeError, ValueError):
                    return False
        if len(str(ws.cell(r, title_col).value)) > 80:
            return False
    return True


def check_final_integrity(ws) -> bool:
    return (
        ws.max_column == 35
        and check_brand_set(ws)
        and check_stock_set(ws)
        and check_video_cleared(ws)
        and check_sku_format(ws)
        and check_image_urls_https(ws)
        and check_duplicate_images(ws)
        and check_required_fields(ws)
    )


CHECKS = {
    "brand_set": check_brand_set,
    "stock_set": check_stock_set,
    "video_cleared": check_video_cleared,
    "sku_format": check_sku_format,
    "image_urls_https": check_image_urls_https,
    "duplicate_images": check_duplicate_images,
    "required_fields": check_required_fields,
    "final_integrity": check_final_integrity,
}


def main(argv: list[str]) -> int:
    if len(argv) != 3 or argv[2] not in CHECKS:
        print(f"usage: check_sheet.py <xlsx> <{'|'.join(CHECKS)}>")
        return 2
    wb = openpyxl.load_workbook(argv[1], data_only=True)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active
    ok = CHECKS[argv[2]](ws)
    print(f"{argv[2]}: {'PASS' if ok else 'FAIL'}")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main(sys.argv))
