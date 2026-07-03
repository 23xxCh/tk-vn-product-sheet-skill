"""Preprocess a raw TikTok product workbook before translation/image work.

1. Deduplicate main/sub-image URLs within each row, preserving first occurrence.
2. Remove description blocks selected by semantic review plus obvious promo noise.
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
from pathlib import Path

import openpyxl


IMAGE_HEADERS = [
    "主图(url)地址", "主图(URL)地址", "产品主图(URL)地址",
    "附图一", "附图二", "附图三", "附图四", "附图五", "附图六", "附图七", "附图八",
]
DESC_HEADERS = ["Tiktok产品描述", "TikTok产品描述", "产品描述"]
DIRTY_TEXT_MARKERS = (
    "联系客服", "客服", "售后", "退换货", "七天无理由", "7天无理由",
    "优惠券", "满减", "关注领券", "收藏有礼", "运费说明", "发货说明",
    "物流时效", "双11", "618", "店铺声明", "认准", "专卖店",
)
IMG_RE = re.compile(r"<img\b[^>]*?\bsrc\s*=\s*(['\"])(.*?)\1[^>]*>", re.I | re.S)
BLOCK_RE = re.compile(r"<(p|div|li)\b[^>]*>.*?</\1>", re.I | re.S)


def _norm(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "")).lower()


def _column_map(ws) -> dict[str, int]:
    return {_norm(ws.cell(1, col).value): col for col in range(1, ws.max_column + 1)}


def _load_manifest(path: str | None) -> dict:
    if not path:
        return {"rows": {}}
    return json.loads(Path(path).read_text(encoding="utf-8"))


def _clean_description(html: str, row_rule: dict) -> tuple[str, int]:
    remove_urls = set(row_rule.get("remove_image_urls", []))
    remove_exact = {re.sub(r"\s+", " ", x).strip() for x in row_rule.get("remove_text_exact", [])}
    removed = 0

    def replace_img(match: re.Match) -> str:
        nonlocal removed
        if match.group(2) in remove_urls:
            removed += 1
            return ""
        return match.group(0)

    html = IMG_RE.sub(replace_img, html)

    def replace_block(match: re.Match) -> str:
        nonlocal removed
        block = match.group(0)
        text = re.sub(r"<[^>]+>", " ", block)
        text = re.sub(r"\s+", " ", text).strip()
        if text in remove_exact or any(marker in text for marker in DIRTY_TEXT_MARKERS):
            removed += 1
            return ""
        return block

    html = BLOCK_RE.sub(replace_block, html)
    return html.strip(), removed


def preprocess(input_path: str, output_path: str, manifest_path: str | None = None) -> dict:
    src = Path(input_path).resolve()
    dst = Path(output_path).resolve()
    if src != dst:
        shutil.copy2(src, dst)

    manifest = _load_manifest(manifest_path)
    wb = openpyxl.load_workbook(dst)
    report = {"rows": 0, "duplicate_image_urls_removed": 0, "description_items_removed": 0}

    for ws in wb.worksheets:
        columns = _column_map(ws)
        image_cols = []
        for header in IMAGE_HEADERS:
            col = columns.get(_norm(header))
            if col and col not in image_cols:
                image_cols.append(col)
        desc_col = next((columns.get(_norm(h)) for h in DESC_HEADERS if columns.get(_norm(h))), None)

        for row in range(2, ws.max_row + 1):
            report["rows"] += 1
            # Preserve main→sub order and compact unique URLs leftward.
            values = [ws.cell(row, col).value for col in image_cols]
            unique, seen = [], set()
            for value in values:
                url = str(value or "").strip()
                if not url:
                    continue
                key = url.lower()
                if key in seen:
                    report["duplicate_image_urls_removed"] += 1
                    continue
                seen.add(key)
                unique.append(url)
            for index, col in enumerate(image_cols):
                ws.cell(row, col).value = unique[index] if index < len(unique) else None

            if desc_col:
                value = ws.cell(row, desc_col).value
                if value:
                    rule = manifest.get("rows", {}).get(str(row), {})
                    cleaned, count = _clean_description(str(value), rule)
                    ws.cell(row, desc_col).value = cleaned
                    report["description_items_removed"] += count

    wb.save(dst)
    return report


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Preprocess raw TikTok product workbook")
    parser.add_argument("xlsx")
    parser.add_argument("-o", "--output", required=True)
    parser.add_argument(
        "--dirty-manifest",
        help="JSON created by semantic review: rows.<row>.remove_image_urls/remove_text_exact",
    )
    args = parser.parse_args(argv)
    print(json.dumps(preprocess(args.xlsx, args.output, args.dirty_manifest), ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
